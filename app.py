from flask import Flask, request, render_template, url_for, redirect, flash, abort, send_file
from flask_login import UserMixin, login_user, LoginManager, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename
from flask_sqlalchemy import SQLAlchemy
from string import ascii_lowercase
from io import BytesIO
import itertools
import datetime
import sqlite3
import random
import string
import json

with open("credenziali.json", "r") as f:
    cr = json.load(f)

with open("costanti.json", "r") as f:
    ct = json.load(f)

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] =  "sqlite:///utenti.db"
app.config['SECRET_KEY'] = cr["secret_key"]

db = SQLAlchemy(app)


class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), nullable=False, unique=True)
    password = db.Column(db.String(128), nullable=False)
    anagrafica = db.Column(db.JSON, nullable=False)
    collaboratore = db.Column(db.JSON, nullable=False)
    verifica = db.Column(db.JSON, nullable=False)

#with app.app_context():
#    db.create_all()

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

utentiDB_path = "utenti.db"
mcpDB_path = "mcp.db"
static_path = cr["static_path"]

#Genera e trova i numeri commessa
def iter_all_strings():
    for size in itertools.count(1):
        for s in itertools.product(ascii_lowercase, repeat=size):
            yield "".join(s)

def genera_int():
    mcpDB = sqlite3.connect(mcpDB_path)
    cur = mcpDB.cursor()
    cur.execute("select * from interne")
    interne = cur.fetchall()
    mcpDB.commit()
    mcpDB.close()
    n = 1
    for i in interne:
        n = trova(i[1]) + 1
    i = 1
    for s in itertools.islice(iter_all_strings(), n):
        if i == n:
            return s.upper()
        i += 1

def trova(n):
    i = 1
    for s in iter_all_strings():
        if s == n.lower():
            return i
        i += 1

def genera_ext():
    mcpDB = sqlite3.connect(mcpDB_path)
    cur = mcpDB.cursor()
    cur.execute("select * from commesse")
    commesse = cur.fetchall()
    mcpDB.commit()
    mcpDB.close()
    n = 1
    for i in commesse:
        n = int(i[1]) + 1
    return n

# ---------------------------

def gestisce():
    gestione = False
    if current_user.collaboratore["ruolo"] == "manager":
        gestione = True
    return gestione

def get_clienti():
    mcpDB = sqlite3.connect(mcpDB_path)
    cur = mcpDB.cursor()
    cur.execute("select * from clienti")
    clienti = cur.fetchall()
    mcpDB.commit()
    mcpDB.close()
    return clienti

def get_tipo_lavorazione():
    mcpDB = sqlite3.connect(mcpDB_path)
    cur = mcpDB.cursor()
    cur.execute("select * from tipo_lavorazioni")
    tipo_lavorazioni = cur.fetchall()
    mcpDB.commit()
    mcpDB.close()
    return tipo_lavorazioni

def get_nome(tipo, tmp_id):
    mcpDB = sqlite3.connect(mcpDB_path)
    cur = mcpDB.cursor()
    if tipo == "cliente":
        cur.execute("select nome from clienti where id = ?", [tmp_id])
        cliente = cur.fetchall()
        try:
            nome = cliente[0][0]
        except:
            nome = ""
    elif tipo == "tipo_lavorazione":
        nome = "Altro"
        if int(tmp_id) != 0:
            cur.execute("select nome from tipo_lavorazioni where id = ?", [tmp_id])
            tipo_lavorazione = cur.fetchall()
            try:
                nome = tipo_lavorazione[0][0]
            except:
                nome = ""
    elif tipo == "collaboratore":
        user = User.query.filter_by(id=tmp_id).first()
        try:
            nome = user.anagrafica["nome"] + " " + user.anagrafica["cognome"]
        except:
            nome = ""
    elif tipo == "file":
        cur.execute("select nome from file where id = ?", [tmp_id])
        tmp_file = cur.fetchall()
        try:
            nome = tmp_file[0][0]
        except:
            nome = ""
    mcpDB.commit()
    mcpDB.close()
    return nome

def get_numero_commessa(numero):
    l_base = 4
    new_numero = ""
    lunghezza = len(numero)
    for i in range(l_base-lunghezza):
        new_numero = "0"+new_numero
    new_numero = new_numero+numero
    return new_numero

def get_numero_offerta(numero, data):
    new_numero = ct["offerte"]["prefisso"]+"-"+numero+ct["offerte"]["suffisso"]+"-"+data["data"][0:4]
    return new_numero

def get_tipo_commessa(lista):
    new_lista = []
    for i in lista:
        tmp_tipo = {"id": i,"nome": get_nome("tipo_lavorazione", i)}
        new_lista.append(tmp_tipo)
    return new_lista

def get_costo(tmp_id, durata):
    costo = 0.0
    tmp_durata = float(durata["ore"])
    if durata["trasferta"]:
        costo += ct["trasferta"]
        tmp_durata = 8.0
    user = User.query.filter_by(id=tmp_id).first()
    user_costo = float(user.collaboratore["costo"])
    costo = costo + (tmp_durata*user_costo)
    return costo

def anni_commesse():
    anni = []
    commesse = dati_mcp("commesse")
    for i in commesse:
        if not str(i["durata"]["inizio"][0:4]) in anni:
            anni.append(str(i["durata"]["inizio"][0:4]))
    return anni

def get_grafico_commessa(dati):
    try:
        guadagno = dati["budget_euro"] - dati["euro_spesi"]
        tmp_ind = int(guadagno/dati["ore_lavorate"])
    except ZeroDivisionError:
        tmp_ind = 28
    grafico = "red"
    if tmp_ind < 23:
        grafico = "red"
    elif tmp_ind > 27:
        grafico = "green"
    else:
        grafico = "yellow"
    return grafico

def insert_edit_lavorazione(dati):
    mcpDB = sqlite3.connect(mcpDB_path)
    cur = mcpDB.cursor()
    if dati["form"] == "inserisci_lavorazione":
        cur.execute("INSERT INTO lavorazioni (collaboratore, commessa, tipo, durata) VALUES (?,?,?,?)", [int(current_user.id), int(dati["id_commessa"]), int(dati["tipo"]), json.dumps(dati["durata"])])
    elif dati["form"] == "modifica_lavorazione":
        cur.execute("UPDATE lavorazioni SET tipo = ?, durata = ? WHERE id = ?", [int(dati["tipo"]), json.dumps(dati["durata"]), int(dati["id_lavorazione"])])
    elif dati["form"] == "elimina_lavorazione":
        cur.execute("DELETE from lavorazioni where id = ?", [int(dati["id_lavorazione"])])
    mcpDB.commit()
    mcpDB.close()

def insert_edit_ore_interne(dati):
    mcpDB = sqlite3.connect(mcpDB_path)
    cur = mcpDB.cursor()
    if dati["form"] == "inserisci_lavorazione":
        cur.execute("INSERT INTO ore_interne (collaboratore, id_interna, durata) VALUES (?,?,?)", [int(current_user.id), int(dati["id_commessa"]), json.dumps(dati["durata"])])
    elif dati["form"] == "modifica_lavorazione":
        cur.execute("UPDATE ore_interne SET durata = ? WHERE id = ?", [json.dumps(dati["durata"]), int(dati["id_lavorazione"])])
    elif dati["form"] == "elimina_lavorazione":
        cur.execute("DELETE from ore_interne where id = ?", [int(dati["id_lavorazione"])])
    mcpDB.commit()
    mcpDB.close()

def insert_edit_commessa(dati):
    mcpDB = sqlite3.connect(mcpDB_path)
    cur = mcpDB.cursor()
    if dati["form"] == "inserisci_commessa":
        specifiche = {
            "budget_ore": int(dati["ore"]),
            "budget_euro": int(dati["euro"]),
            "project_manager": int(dati["project_manager"]),
            "tipologia": dati["tipologia"]
            }
        durata = {
            "inizio": dati["data"],
            "fine": False
            }
        numero = str(genera_ext())
        if dati["id_offerta"] == "":
            tmp_id_offerta = 0
        else:
            tmp_id_offerta = int(dati["id_offerta"])
        cur.execute("INSERT INTO commesse (numero, nome, cliente, specifiche, durata, note, offerta, file) VALUES (?,?,?,?,?,?,?,?)", [numero, dati["nome"], int(dati["cliente"]), json.dumps(specifiche), json.dumps(durata), dati["note"], tmp_id_offerta, json.dumps({})])
    elif dati["form"] == "modifica_commessa":
        specifiche = {
            "budget_ore": int(dati["ore"]),
            "budget_euro": int(dati["euro"]),
            "project_manager": int(dati["project_manager"]),
            "tipologia": dati["tipologia"]
            }
        durata = {
            "inizio": dati["data"],
            "fine": False
            }
        cur.execute("UPDATE commesse SET nome = ?, cliente = ?, specifiche = ?, durata = ?, note = ? WHERE id = ?", [dati["nome"], int(dati["cliente"]), json.dumps(specifiche), json.dumps(durata), dati["note"], dati["id_commessa"]])
    elif dati["form"] == "elimina_commessa":
        cur.execute("DELETE from commesse where id = ?", [int(dati["id_commessa"])])
        cur.execute("DELETE from lavorazioni where commessa = ?", [int(dati["id_commessa"])])
    elif dati["form"] == "chiudi_commessa":
        cur.execute("select durata from commesse where id = ?", [dati["id_commessa"]])
        commessa = cur.fetchall()
        tmp_durata = json.loads(commessa[0][0])
        durata = {
            "inizio": tmp_durata["inizio"],
            "fine": dati["data"]
            }
        cur.execute("UPDATE commesse SET durata = ? WHERE id = ?", [json.dumps(durata), dati["id_commessa"]])
    mcpDB.commit()
    mcpDB.close()

def insert_edit_offerta(dati):
    mcpDB = sqlite3.connect(mcpDB_path)
    cur = mcpDB.cursor()
    if dati["form"] == "inserisci_offerta":
        versioni = {
            "data": dati["data"],
            "chiusa": False,
            "versioni": []
            }
        cur.execute("INSERT INTO offerte (numero, ordine_rif, cliente, versioni, note) VALUES (?,?,?,?,?)", [dati["nome"], dati["n_ordine"], int(dati["cliente"]), json.dumps(versioni), dati["note"]])
    elif dati["form"] == "modifica_offerta":
        cur.execute("UPDATE offerte SET numero = ?, ordine_rif = ?, cliente = ?, note = ? WHERE id = ?", [dati["nome"], dati["n_ordine"], int(dati["cliente"]), dati["note"], int(dati["id_offerta"])])
    elif dati["form"] == "elimina_offerta":
        cur.execute("DELETE from offerte where id = ?", [int(dati["id_offerta"])])
    mcpDB.commit()
    mcpDB.close()

def insert_edit_commessa_interna(dati):
    mcpDB = sqlite3.connect(mcpDB_path)
    cur = mcpDB.cursor()
    if dati["form"] == "inserisci_commessa":
        durata = {
            "inizio": dati["data"],
            "fine": False
            }
        numero = str(genera_int())
        cur.execute("INSERT INTO interne (numero, nome, durata, note) VALUES (?,?,?,?)", [numero, dati["nome"], json.dumps(durata), dati["note"]])
    elif dati["form"] == "modifica_commessa":
        durata = {
            "inizio": dati["data"],
            "fine": False
            }
        cur.execute("UPDATE interne SET nome = ?, durata = ?, note = ? WHERE id = ?", [dati["nome"], json.dumps(durata), dati["note"], dati["id_commessa"]])
    elif dati["form"] == "elimina_commessa":
        cur.execute("DELETE from interne where id = ?", [int(dati["id_commessa"])])
        cur.execute("DELETE from ore_interne where id_interna = ?", [int(dati["id_commessa"])])
    elif dati["form"] == "chiudi_commessa":
        cur.execute("select durata from interne where id = ?", [dati["id_commessa"]])
        commessa = cur.fetchall()
        tmp_durata = json.loads(commessa[0][0])
        durata = {
            "inizio": tmp_durata["inizio"],
            "fine": dati["data"]
            }
        cur.execute("UPDATE interne SET durata = ? WHERE id = ?", [json.dumps(durata), dati["id_commessa"]])
    mcpDB.commit()
    mcpDB.close()

def dati_mcp(tipo):
    mcpDB = sqlite3.connect(mcpDB_path)
    cur = mcpDB.cursor()

    if tipo == "commesse":
        cur.execute("select * from commesse")
        commesse = cur.fetchall()
        dati = []
        for commessa in commesse:
            tmp_specifiche = {
                    "budget_ore": json.loads(commessa[4])["budget_ore"],
                    "budget_euro": json.loads(commessa[4])["budget_euro"],
                    "project_manager": {"id": json.loads(commessa[4])["project_manager"],"nome": get_nome("collaboratore", json.loads(commessa[4])["project_manager"])},
                    "tipologia": get_tipo_commessa(json.loads(commessa[4])["tipologia"])
                    }
            if commessa[7] != 0:
                cur.execute("select numero, versioni from offerte where id = ?", [commessa[7]])
                offerte = cur.fetchall()[0]
                tmp_id_offerta = int(commessa[7])
                tmp_numero = get_numero_offerta(offerte[0], json.loads(offerte[1]))
            else:
                tmp_id_offerta = 0
                tmp_numero = 0
            tmp_commessa = {
                "id": commessa[0],
                "numero": get_numero_commessa(commessa[1]),
                "nome": commessa[2],
                "cliente": {"id": commessa[3],"nome": get_nome("cliente", commessa[3])},
                "specifiche": tmp_specifiche,
                "durata": json.loads(commessa[5]),
                "note": commessa[6],
                "offerta": {"id": tmp_id_offerta,"numero": tmp_numero},
                "file": commessa[8]
                }
            dati.append(tmp_commessa)

    elif tipo == "commesse_complete":
        cur.execute("select * from commesse")
        commesse = cur.fetchall()
        dati = []
        for commessa in commesse:
            cur.execute("select * from lavorazioni where commessa = ?", [commessa[0]])
            lavorazioni = cur.fetchall()
            tmp_lavorazioni = []
            ore_lavorate = 0.0
            euro_spesi = 0.0
            for lavorazione in lavorazioni:
                tmp_lavorazione = {
                    "id": lavorazione[0],
                    "collaboratore": {"id": lavorazione[1],"nome": get_nome("collaboratore", lavorazione[1])},
                    "tipo": {"id": lavorazione[3],"nome": get_nome("tipo_lavorazione", lavorazione[3])},
                    "durata": json.loads(lavorazione[4]),
                    "costo": get_costo(lavorazione[1], json.loads(lavorazione[4]))
                    }
                ore_lavorate += float(tmp_lavorazione["durata"]["ore"])
                euro_spesi += tmp_lavorazione["costo"]
                tmp_lavorazioni.append(tmp_lavorazione)
            tmp_lavorazioni.sort(key = lambda x: datetime.datetime.strptime(x["durata"]["data"], '%Y-%m-%d'))
            tmp_specifiche = {
                    "budget_ore": float(json.loads(commessa[4])["budget_ore"]),
                    "budget_euro": float(json.loads(commessa[4])["budget_euro"]),
                    "ore_lavorate": ore_lavorate,
                    "euro_spesi": euro_spesi,
                    "project_manager": {"id": json.loads(commessa[4])["project_manager"],"nome": get_nome("collaboratore", json.loads(commessa[4])["project_manager"])},
                    "tipologia": get_tipo_commessa(json.loads(commessa[4])["tipologia"])
                    }
            if commessa[7] != 0:
                cur.execute("select numero, versioni from offerte where id = ?", [commessa[7]])
                offerte = cur.fetchall()[0]
                tmp_id_offerta = int(commessa[7])
                tmp_numero = get_numero_offerta(offerte[0], json.loads(offerte[1]))
            else:
                tmp_id_offerta = 0
                tmp_numero = 0
            tmp_commessa = {
                "id": commessa[0],
                "numero": get_numero_commessa(commessa[1]),
                "nome": commessa[2],
                "cliente": {"id": commessa[3],"nome": get_nome("cliente", commessa[3])},
                "specifiche": tmp_specifiche,
                "durata": json.loads(commessa[5]),
                "note": commessa[6],
                "offerta": {"id": tmp_id_offerta,"numero": tmp_numero},
                "file": commessa[8],
                "lavorazioni": tmp_lavorazioni,
                "qualita": get_grafico_commessa(tmp_specifiche)
                }
            dati.append(tmp_commessa)

    elif tipo == "offerte":
        cur.execute("select * from offerte")
        offerte = cur.fetchall()
        dati = []
        for offerta in offerte:
            tmp_versioni = []
            for i in json.loads(offerta[4])["versioni"]:
                tmp_versioni.append({"rev": i["rev"], "file": {"id": i["file"],"nome": get_nome("file", i["file"])}})
            tmp_offerta = {
                "id": offerta[0],
                "numero": get_numero_offerta(offerta[1], json.loads(offerta[4])),
                "ordine_rif": offerta[2],
                "cliente": {"id": offerta[3],"nome": get_nome("cliente", offerta[3])},
                "versioni": {"data": json.loads(offerta[4])["data"], "chiusa": json.loads(offerta[4])["chiusa"], "versioni": tmp_versioni},
                "note": offerta[5]
                }
            dati.append(tmp_offerta)
    elif tipo == "interne":
        cur.execute("select * from interne")
        interne = cur.fetchall()
        dati = []
        for commessa in interne:
            cur.execute("select * from ore_interne where id_interna = ?", [commessa[0]])
            lavorazioni = cur.fetchall()
            tmp_lavorazioni = []
            ore_lavorate = 0.0
            euro_spesi = 0.0
            for lavorazione in lavorazioni:
                tmp_lavorazione = {
                    "id": lavorazione[0],
                    "collaboratore": {"id": lavorazione[1],"nome": get_nome("collaboratore", lavorazione[1])},
                    "durata": json.loads(lavorazione[3]),
                    "costo": get_costo(lavorazione[1], json.loads(lavorazione[3]))
                    }
                ore_lavorate += float(tmp_lavorazione["durata"]["ore"])
                euro_spesi += tmp_lavorazione["costo"]
                tmp_lavorazioni.append(tmp_lavorazione)
            tmp_lavorazioni.sort(key = lambda x: datetime.datetime.strptime(x["durata"]["data"], '%Y-%m-%d'))
            tmp_specifiche = {
                    "ore_lavorate": ore_lavorate,
                    "euro_spesi": euro_spesi,
                    }
            tmp_commessa = {
                "id": commessa[0],
                "numero": get_numero_commessa(commessa[1]),
                "nome": commessa[2],
                "specifiche": tmp_specifiche,
                "durata": json.loads(commessa[3]),
                "note": commessa[4],
                "lavorazioni": tmp_lavorazioni
                }
            dati.append(tmp_commessa)

    mcpDB.commit()
    mcpDB.close()
    return dati

@app.route("/")
def homepage():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    return render_template("index.html")

@app.route("/dashboard")
@login_required
def dashboard():
    return render_template("dashboard.html", gestione=gestisce(), menu_page="dashboard")

@app.route("/commesse", defaults={ "pagina": "riepilogo" }, methods=("GET", "POST"))
@app.route("/commesse/<string:pagina>", methods=("GET", "POST"))
@login_required
def commesse(pagina):
    if request.method == "POST":
        if request.form["id_form"] == "inserisci_lavorazione" or request.form["id_form"] == "modifica_lavorazione":
            try:
                request.form["trasferta"]
                tmp_trasferta = True
            except:
                tmp_trasferta = False
            try:
                anno, mese, giorno = request.form["data"].split('-')
                tmp_ore = float(request.form["ore"])
                ore, minuti = str(tmp_ore).split('.')
                tmp = datetime.time(hour=int(ore), minute=(int(60*(int(minuti)/10))))
                dati_validi = True
            except:
                dati_validi = False
            tmp_date = datetime.datetime.strptime(request.form["data"], "%Y-%m-%d").date()
            start_date = datetime.datetime.now().date() - datetime.timedelta(days=10)
            if tmp_date < start_date and not gestisce():
                dati_validi = False
            if tmp_date > datetime.datetime.now().date():
                flash("Ehi stai viaggiando nel tempo?", "warning")
                dati_validi = False
            tmp_dati = {
                "form": request.form["id_form"],
                "id_commessa": request.form["id_commessa"],
                "id_lavorazione": request.form["id_lavorazione"],
                "tipo": request.form["tipo"],
                "durata": {
                    "data": request.form["data"],
                    "ore": request.form["ore"],
                    "trasferta": tmp_trasferta
                    }
                }
            if dati_validi:
                insert_edit_lavorazione(tmp_dati)
                flash("Inserimento avvenuto con successo!", "success")
            else:
                flash("Dati errati.", "warning")
            pagina = tmp_dati["id_commessa"]
        elif request.form["id_form"] == "elimina_lavorazione":
            tmp_dati = {
                "form": request.form["id_form"],
                "id_commessa": request.form["id_commessa"],
                "id_lavorazione": request.form["id_lavorazione"]
                }
            insert_edit_lavorazione(tmp_dati)
            flash("Eliminato con successo!", "success")
            pagina = tmp_dati["id_commessa"]
        elif request.form["id_form"] == "modifica_commessa":
            tipi = []
            for i in get_tipo_lavorazione():
                try:
                    tipi.append(int(request.form[i[1]]))
                except:
                    pass
            try:
                tipi.append(int(request.form["tipo_altro"]))
            except:
                pass
            try:
                anno, mese, giorno = request.form["data"].split('-')
                tmp = datetime.date(int(anno), int(mese), int(giorno))
                dati_validi = True
            except:
                dati_validi = False
                flash("Dati errati: Manca la data di apertura commessa!", "warning")
            try:
                tmp_ore = int(request.form["ore"])
            except:
                tmp_ore = 0
            try:
                tmp_euro = int(request.form["euro"])
            except:
                tmp_euro = 0
            tmp_dati = {
                "form": request.form["id_form"],
                "id_commessa": request.form["id_commessa"],
                "nome": request.form["nome"],
                "cliente": request.form["cliente"],
                "project_manager": request.form["project_manager"],
                "tipologia": tipi,
                "data": request.form["data"],
                "ore": tmp_ore,
                "euro": tmp_euro,
                "note": request.form["note"]
                }
            mcpDB = sqlite3.connect(mcpDB_path)
            cur = mcpDB.cursor()
            cur.execute("select * from commesse where id = ?", [int(request.form["id_commessa"])])
            commesse = cur.fetchall()
            mcpDB.commit()
            mcpDB.close()
            if dati_validi:
                insert_edit_commessa(tmp_dati)
                flash("Commessa modificata con successo!", "success")
            pagina = tmp_dati["id_commessa"]
        elif request.form["id_form"] == "elimina_commessa":
            tmp_dati = {
                "form": request.form["id_form"],
                "id_commessa": request.form["id_commessa"]
                }
            insert_edit_commessa(tmp_dati)
            flash("Commessa eliminata con successo!", "success")
        elif request.form["id_form"] == "chiudi_commessa":
            tmp_dati = {
                "form": request.form["id_form"],
                "id_commessa": request.form["id_commessa"],
                "data": request.form["data"]
                }
            insert_edit_commessa(tmp_dati)
            flash("Commessa chiusa con successo!", "success")
    if pagina == "riepilogo":
        return render_template("commesse.html", pagina=pagina, gestione=gestisce(), dati=dati_mcp("commesse_complete"), clienti=get_clienti(), utenti=User.query.all(), tipi_lavorazione=get_tipo_lavorazione(), menu_page="commesse")
    # else
    mcpDB = sqlite3.connect(mcpDB_path)
    cur = mcpDB.cursor()
    cur.execute("select * from commesse where id = ?", [int(pagina)])
    commesse = cur.fetchall()
    cur.execute("select * from lavorazioni where commessa = ?", [int(pagina)])
    lavorazioni = cur.fetchall()
    mcpDB.commit()
    mcpDB.close()
    tmp_lavorazioni = []
    ore_lavorate = 0.0
    euro_spesi = 0.0
    for lavorazione in lavorazioni:
        tmp_lavorazione = {
            "id": lavorazione[0],
            "collaboratore": {"id": lavorazione[1],"nome": get_nome("collaboratore", lavorazione[1])},
            "tipo": {"id": lavorazione[3],"nome": get_nome("tipo_lavorazione", lavorazione[3])},
            "durata": json.loads(lavorazione[4]),
            "costo": get_costo(lavorazione[1], json.loads(lavorazione[4]))
            }
        ore_lavorate += float(tmp_lavorazione["durata"]["ore"])
        euro_spesi += tmp_lavorazione["costo"]
        tmp_lavorazioni.append(tmp_lavorazione)
    tmp_lavorazioni.sort(key = lambda x: datetime.datetime.strptime(x["durata"]["data"], '%Y-%m-%d'))
    tmp_specifiche = {
            "budget_ore": float(json.loads(commesse[0][4])["budget_ore"]),
            "budget_euro": float(json.loads(commesse[0][4])["budget_euro"]),
            "ore_lavorate": ore_lavorate,
            "euro_spesi": euro_spesi,
            "project_manager": {"id": json.loads(commesse[0][4])["project_manager"],"nome": get_nome("collaboratore", json.loads(commesse[0][4])["project_manager"])},
            "tipologia": get_tipo_commessa(json.loads(commesse[0][4])["tipologia"])
            }
    if commesse[0][7] != 0:
        mcpDB = sqlite3.connect(mcpDB_path)
        cur = mcpDB.cursor()
        cur.execute("select numero, versioni from offerte where id = ?", [commesse[0][7]])
        offerte = cur.fetchall()[0]
        tmp_id_offerta = int(commesse[0][7])
        tmp_numero = get_numero_offerta(offerte[0], json.loads(offerte[1]))
        mcpDB.commit()
        mcpDB.close()
    else:
        tmp_id_offerta = 0
        tmp_numero = 0
    tmp_commessa = {
        "id": commesse[0][0],
        "numero": get_numero_commessa(commesse[0][1]),
        "nome": commesse[0][2],
        "cliente": {"id": commesse[0][3],"nome": get_nome("cliente", commesse[0][3])},
        "specifiche": tmp_specifiche,
        "durata": json.loads(commesse[0][5]),
        "note": commesse[0][6],
        "lavorazioni": tmp_lavorazioni,
        "offerta": {"id": tmp_id_offerta,"numero": tmp_numero},
        "qualita": get_grafico_commessa(tmp_specifiche)
        }
    return render_template("dettaglio_commessa.html", commessa=tmp_commessa, gestione=gestisce(), clienti=get_clienti(), utenti=User.query.all(), tipi_lavorazione=get_tipo_lavorazione(), menu_page="commesse")

@app.route("/offerte", methods=("GET", "POST"))
@login_required
def offerte():
    return render_template("offerte.html", gestione=gestisce(), dati=reversed(dati_mcp("offerte")), clienti=get_clienti(), utenti=User.query.all(), menu_page="offerte")

@app.route("/interne", defaults={ "pagina": "riepilogo" }, methods=("GET", "POST"))
@app.route("/interne/<string:pagina>", methods=("GET", "POST"))
@login_required
def interne(pagina):
    if request.method == "POST":
        if request.form["id_form"] == "inserisci_lavorazione" or request.form["id_form"] == "modifica_lavorazione":
            try:
                request.form["trasferta"]
                tmp_trasferta = True
            except:
                tmp_trasferta = False
            try:
                anno, mese, giorno = request.form["data"].split('-')
                tmp_ore = float(request.form["ore"])
                ore, minuti = str(tmp_ore).split('.')
                tmp = datetime.time(hour=int(ore), minute=(int(60*(int(minuti)/10))))
                dati_validi = True
            except:
                dati_validi = False
            tmp_date = datetime.datetime.strptime(request.form["data"], "%Y-%m-%d").date()
            start_date = datetime.datetime.now().date() - datetime.timedelta(days=10)
            if tmp_date < start_date and not gestisce():
                dati_validi = False
            if tmp_date > datetime.datetime.now().date():
                flash("Ehi stai viaggiando nel tempo?", "warning")
                dati_validi = False
            tmp_dati = {
                "form": request.form["id_form"],
                "id_commessa": request.form["id_commessa"],
                "id_lavorazione": request.form["id_lavorazione"],
                "durata": {
                    "data": request.form["data"],
                    "ore": request.form["ore"],
                    "trasferta": tmp_trasferta
                    }
                }
            if dati_validi:
                insert_edit_ore_interne(tmp_dati)
                flash("Inserimento avvenuto con successo!", "success")
            else:
                flash("Dati errati.", "warning")
            pagina = tmp_dati["id_commessa"]
        elif request.form["id_form"] == "elimina_lavorazione":
            tmp_dati = {
                "form": request.form["id_form"],
                "id_commessa": request.form["id_commessa"],
                "id_lavorazione": request.form["id_lavorazione"]
                }
            insert_edit_ore_interne(tmp_dati)
            flash("Eliminato con successo!", "success")
            pagina = tmp_dati["id_commessa"]
        elif request.form["id_form"] == "modifica_commessa":
            try:
                anno, mese, giorno = request.form["data"].split('-')
                tmp = datetime.date(int(anno), int(mese), int(giorno))
                dati_validi = True
            except:
                dati_validi = False
                flash("Dati errati: Manca la data di apertura commessa!", "warning")
            tmp_date = datetime.datetime.strptime(request.form["data"], "%Y-%m-%d").date()
            start_date = datetime.datetime.now().date() - datetime.timedelta(days=10)
            if tmp_date > datetime.datetime.now().date():
                flash("Ehi stai viaggiando nel tempo?", "warning")
            tmp_dati = {
                "form": request.form["id_form"],
                "id_commessa": request.form["id_commessa"],
                "nome": request.form["nome"],
                "data": request.form["data"],
                "note": request.form["note"]
                }
            mcpDB = sqlite3.connect(mcpDB_path)
            cur = mcpDB.cursor()
            cur.execute("select * from commesse where id = ?", [int(request.form["id_commessa"])])
            commesse = cur.fetchall()
            mcpDB.commit()
            mcpDB.close()
            if dati_validi:
                insert_edit_commessa_interna(tmp_dati)
                flash("Commessa modificata con successo!", "success")
            pagina = tmp_dati["id_commessa"]
        elif request.form["id_form"] == "elimina_commessa":
            tmp_dati = {
                "form": request.form["id_form"],
                "id_commessa": request.form["id_commessa"]
                }
            insert_edit_commessa_interna(tmp_dati)
            flash("Commessa eliminata con successo!", "success")
        elif request.form["id_form"] == "chiudi_commessa":
            tmp_dati = {
                "form": request.form["id_form"],
                "id_commessa": request.form["id_commessa"],
                "data": request.form["data"]
                }
            insert_edit_commessa_interna(tmp_dati)
            flash("Commessa chiusa con successo!", "success")
    return render_template("interne.html", pagina=pagina, gestione=gestisce(), dati=dati_mcp("interne"), clienti=get_clienti(), utenti=User.query.all(), menu_page="interne")

@app.route("/crea_commessa", defaults={ "offerta": False }, methods=("GET", "POST"))
@app.route("/crea_commessa/<string:offerta>", methods=("GET", "POST"))
@login_required
def crea_commessa(offerta):
    if request.method == "POST":
        if request.form["id_form"] == "inserisci_commessa":
            tipi = []
            for i in get_tipo_lavorazione():
                try:
                    tipi.append(int(request.form[i[1]]))
                except:
                    pass
            try:
                tipi.append(int(request.form["tipo_altro"]))
            except:
                pass
            try:
                anno, mese, giorno = request.form["data"].split('-')
                tmp = datetime.date(int(anno), int(mese), int(giorno))
                dati_validi = True
            except:
                dati_validi = False
                flash("Dati errati: Manca la data di apertura commessa!", "warning")
            tmp_date = datetime.datetime.strptime(request.form["data"], "%Y-%m-%d").date()
            if tmp_date > datetime.datetime.now().date():
                flash("Ehi stai viaggiando nel tempo?", "warning")
                dati_validi = False
            try:
                tmp_ore = int(request.form["ore"])
            except:
                tmp_ore = 0
            try:
                tmp_euro = int(request.form["euro"])
            except:
                tmp_euro = 0
            tmp_id_cliente = request.form["cliente"]
            if request.form["nome"] == "":
                dati_validi = False
                flash("Dati errati: Non puoi creare una commessa senza nome!", "warning")
            if request.form["cliente"] == "0":
                if request.form["new_cliente"] != "" and dati_validi:
                    mcpDB = sqlite3.connect(mcpDB_path)
                    cur = mcpDB.cursor()
                    cur.execute("INSERT INTO clienti (nome) VALUES (?)", [request.form["new_cliente"]])
                    cur.execute("select * from clienti where nome = ?", [request.form["new_cliente"]])
                    tmp_clienti = cur.fetchall()
                    mcpDB.commit()
                    mcpDB.close()
                    tmp_id_cliente = tmp_clienti[0][0]
                else:
                    dati_validi = False
                    flash("Dati errati: Non puoi aggiungere un nuovo cliente senza nome!", "warning")

            tmp_dati = {
                "form": request.form["id_form"],
                "id_commessa": "",
                "id_offerta": request.form["id_offerta"],
                "nome": request.form["nome"],
                "cliente": tmp_id_cliente,
                "project_manager": request.form["project_manager"],
                "tipologia": tipi,
                "data": request.form["data"],
                "ore": tmp_ore,
                "euro": tmp_euro,
                "note": request.form["note"]
                }
            if dati_validi:
                insert_edit_commessa(tmp_dati)
                flash("Inserito con successo!", "success")
        return redirect(url_for("commesse"))
    if offerta:
        mcpDB = sqlite3.connect(mcpDB_path)
        cur = mcpDB.cursor()
        cur.execute("select * from offerte where id = ?", [offerta])
        offerte = cur.fetchall()
        for offerta in offerte:
            tmp_versioni = []
            for i in json.loads(offerta[4])["versioni"]:
                tmp_versioni.append({"rev": i["rev"], "file": {"id": i["file"],"nome": get_nome("file", i["file"])}})
            tmp_offerta = {
                "id": offerta[0],
                "numero": get_numero_offerta(offerta[1], json.loads(offerta[4])),
                "ordine_rif": offerta[2],
                "cliente": {"id": offerta[3],"nome": get_nome("cliente", offerta[3])},
                "versioni": {"data": json.loads(offerta[4])["data"], "chiusa": json.loads(offerta[4])["chiusa"], "versioni": tmp_versioni},
                "note": offerta[5]
                }
        mcpDB.commit()
        mcpDB.close()
        return render_template("crea_commessa_da_offerta.html", gestione=gestisce(), clienti=get_clienti(), utenti=User.query.all(), tipi_lavorazione=get_tipo_lavorazione(), offerta=tmp_offerta, menu_page="crea_commessa")
    return render_template("crea_commessa.html", gestione=gestisce(), clienti=get_clienti(), utenti=User.query.all(), tipi_lavorazione=get_tipo_lavorazione(), menu_page="crea_commessa")


@app.route("/crea_offerta", methods=("GET", "POST"))
@login_required
def crea_offerta():
    if request.method == "POST":
        if request.form["id_form"] == "inserisci_offerta":
            try:
                anno, mese, giorno = request.form["data"].split('-')
                tmp = datetime.date(int(anno), int(mese), int(giorno))
                dati_validi = True
            except:
                dati_validi = False
                flash("Dati errati: Manca la data di creazione offerta!", "warning")
            tmp_date = datetime.datetime.strptime(request.form["data"], "%Y-%m-%d").date()
            if tmp_date > datetime.datetime.now().date():
                flash("Ehi stai viaggiando nel tempo?", "warning")
                dati_validi = False
            tmp_id_cliente = request.form["cliente"]
            if request.form["nome"] == "":
                dati_validi = False
                flash("Dati errati: Non puoi creare un'offerta senza ID!", "warning")
            if request.form["n_ordine"] == "":
                n_ordine = ""
            else:
                n_ordine = request.form["n_ordine"]
            if request.form["cliente"] == "0":
                if request.form["new_cliente"] != "" and dati_validi:
                    mcpDB = sqlite3.connect(mcpDB_path)
                    cur = mcpDB.cursor()
                    cur.execute("INSERT INTO clienti (nome) VALUES (?)", [request.form["new_cliente"]])
                    cur.execute("select * from clienti where nome = ?", [request.form["new_cliente"]])
                    tmp_clienti = cur.fetchall()
                    mcpDB.commit()
                    mcpDB.close()
                    tmp_id_cliente = tmp_clienti[0][0]
                else:
                    dati_validi = False
                    flash("Dati errati: Non puoi aggiungere un nuovo cliente senza nome!", "warning")

            tmp_dati = {
                "form": request.form["id_form"],
                "nome": request.form["nome"],
                "cliente": tmp_id_cliente,
                "n_ordine": n_ordine,
                "data": request.form["data"],
                "note": request.form["note"]
                }
            if dati_validi:
                insert_edit_offerta(tmp_dati)
                flash("Inserito con successo!", "success")
        return redirect(url_for("offerte"))
    return render_template("crea_offerta.html", anno=datetime.datetime.now().year, gestione=gestisce(), clienti=get_clienti(), utenti=User.query.all(), pre_suff=ct["offerte"], menu_page="crea_offerta")

@app.route("/crea_interna", methods=("GET", "POST"))
@login_required
def crea_interna():
    if request.method == "POST":
        if request.form["id_form"] == "inserisci_commessa":
            try:
                anno, mese, giorno = request.form["data"].split('-')
                tmp = datetime.date(int(anno), int(mese), int(giorno))
                dati_validi = True
            except:
                dati_validi = False
                flash("Dati errati: Manca la data di apertura commessa!", "warning")
            tmp_date = datetime.datetime.strptime(request.form["data"], "%Y-%m-%d").date()
            if tmp_date > datetime.datetime.now().date():
                flash("Ehi stai viaggiando nel tempo?", "warning")
                dati_validi = False
            if request.form["nome"] == "":
                dati_validi = False
                flash("Dati errati: Non puoi creare una commessa senza nome!", "warning")
            tmp_dati = {
                "form": request.form["id_form"],
                "id_commessa": "",
                "nome": request.form["nome"],
                "data": request.form["data"],
                "note": request.form["note"]
                }
            if dati_validi:
                insert_edit_commessa_interna(tmp_dati)
                flash("Inserito con successo!", "success")
        return redirect(url_for("interne"))
    return render_template("crea_interna.html", gestione=gestisce(), clienti=get_clienti(), utenti=User.query.all(), tipi_lavorazione=get_tipo_lavorazione(), menu_page="crea_interna")

@app.route("/dettaglio_commessa/<string:id_commessa>")
@login_required
def dettaglio_storico(id_commessa):
    if not gestisce():
        return redirect(url_for("dashboard"))
    mcpDB = sqlite3.connect(mcpDB_path)
    cur = mcpDB.cursor()
    cur.execute("select * from commesse where id = ?", [int(id_commessa)])
    commesse = cur.fetchall()
    cur.execute("select * from lavorazioni where commessa = ?", [int(id_commessa)])
    lavorazioni = cur.fetchall()
    mcpDB.commit()
    mcpDB.close()
    tmp_lavorazioni = []
    ore_lavorate = 0.0
    euro_spesi = 0.0
    for lavorazione in lavorazioni:
        tmp_lavorazione = {
            "id": lavorazione[0],
            "collaboratore": {"id": lavorazione[1],"nome": get_nome("collaboratore", lavorazione[1])},
            "tipo": {"id": lavorazione[3],"nome": get_nome("tipo_lavorazione", lavorazione[3])},
            "durata": json.loads(lavorazione[4]),
            "costo": get_costo(lavorazione[1], json.loads(lavorazione[4]))
            }
        ore_lavorate += float(tmp_lavorazione["durata"]["ore"])
        euro_spesi += tmp_lavorazione["costo"]
        tmp_lavorazioni.append(tmp_lavorazione)
    tmp_lavorazioni.sort(key = lambda x: datetime.datetime.strptime(x["durata"]["data"], '%Y-%m-%d'))
    tmp_specifiche = {
            "budget_ore": float(json.loads(commesse[0][4])["budget_ore"]),
            "budget_euro": float(json.loads(commesse[0][4])["budget_euro"]),
            "ore_lavorate": ore_lavorate,
            "euro_spesi": euro_spesi,
            "project_manager": {"id": json.loads(commesse[0][4])["project_manager"],"nome": get_nome("collaboratore", json.loads(commesse[0][4])["project_manager"])},
            "tipologia": get_tipo_commessa(json.loads(commesse[0][4])["tipologia"])
            }
    if commesse[0][7] != 0:
        mcpDB = sqlite3.connect(mcpDB_path)
        cur = mcpDB.cursor()
        cur.execute("select numero, versioni from offerte where id = ?", [commesse[0][7]])
        offerte = cur.fetchall()[0]
        tmp_id_offerta = int(commesse[0][7])
        tmp_numero = get_numero_offerta(offerte[0], json.loads(offerte[1]))
        mcpDB.commit()
        mcpDB.close()
    else:
        tmp_id_offerta = 0
        tmp_numero = 0
    tmp_commessa = {
        "id": commesse[0][0],
        "numero": get_numero_commessa(commesse[0][1]),
        "nome": commesse[0][2],
        "cliente": {"id": commesse[0][3],"nome": get_nome("cliente", commesse[0][3])},
        "specifiche": tmp_specifiche,
        "durata": json.loads(commesse[0][5]),
        "note": commesse[0][6],
        "lavorazioni": tmp_lavorazioni,
        "offerta": {"id": tmp_id_offerta,"numero": tmp_numero},
        "qualita": get_grafico_commessa(tmp_specifiche)
        }
    return render_template("dettaglio_storico.html", commessa=tmp_commessa, gestione=gestisce(), clienti=get_clienti(), utenti=User.query.all(), tipi_lavorazione=get_tipo_lavorazione(), menu_page="storico")

@app.route("/dettaglio_offerta/<string:id_offerta>", methods=("GET", "POST"))
@login_required
def dettaglio_offerta(id_offerta):
    if request.method == "POST":
        if request.form["id_form"] == "inserisci_file":
            try:
                mcpDB = sqlite3.connect(mcpDB_path)
                cur = mcpDB.cursor()
                file = request.files["file"]
                if file.filename == "":
                    flash("Nessun file selezionato!", "warning")
                    return render_template("offerte.html", pagina=pagina, gestione=gestisce(), dati=dati_mcp("offerte"), clienti=get_clienti(), utenti=User.query.all(), menu_page="offerte")
                cur.execute("INSERT INTO file (nome, dati) VALUES (?,?)", [secure_filename(file.filename), file.read()])
                id_file = cur.lastrowid
                cur.execute("select versioni from offerte where id = ?", [int(request.form["id_offerta"])])
                offerte = cur.fetchall()
                tmp_rev = 0
                tmp_versioni = json.loads(offerte[0][0])
                if len(tmp_versioni["versioni"]) > 0:
                    for i in tmp_versioni["versioni"]:
                        if i["rev"] > tmp_rev:
                            tmp_rev = i["rev"]
                    tmp_rev += 1
                tmp_versioni["versioni"].append({"rev": tmp_rev,"file": id_file})

                cur.execute("UPDATE offerte SET versioni = ? WHERE id = ?", [json.dumps(tmp_versioni), int(request.form["id_offerta"])])

                mcpDB.commit()
                mcpDB.close()
                flash("File caricato con successo!", "success")
            except:
                flash("Errore nel caricamento del file!", "warning")
        elif request.form["id_form"] == "aggiorna_file":
            try:
                mcpDB = sqlite3.connect(mcpDB_path)
                cur = mcpDB.cursor()
                file = request.files["file"]
                if file.filename == "":
                    flash("Nessun file selezionato!", "warning")
                    return render_template("offerte.html", pagina=pagina, gestione=gestisce(), dati=dati_mcp("offerte"), clienti=get_clienti(), utenti=User.query.all(), menu_page="offerte")
                tmp_uuid = str(uuid.uuid4()) + str(datetime.datetime.now()).replace(" ", "-")
                cur.execute("UPDATE file SET nome = ?, dati = ?, uuid = ? WHERE id = ?", [secure_filename(file.filename), file.read(), tmp_uuid, int(request.form["id_file"])])
                mcpDB.commit()
                mcpDB.close()
                flash("File caricato con successo!", "success")
            except:
                flash("Errore nel caricamento del file!", "warning")
        elif request.form["id_form"] == "elimina_file":
            try:
                mcpDB = sqlite3.connect(mcpDB_path)
                cur = mcpDB.cursor()
                cur.execute("DELETE from file where id = ?", [int(request.form["id_file"])])
                cur.execute("select versioni from offerte where id = ?", [int(request.form["id_offerta"])])
                offerte = cur.fetchall()
                tmp_versioni = json.loads(offerte[0][0])
                tmp_tmp_versioni = {"data": tmp_versioni["data"], "chiusa": tmp_versioni["chiusa"], "versioni": []}
                for i in tmp_versioni["versioni"]:
                    if i["file"] != int(request.form["id_file"]):
                        tmp_tmp_versioni["versioni"].append(i)
                cur.execute("UPDATE offerte SET versioni = ? WHERE id = ?", [json.dumps(tmp_tmp_versioni), int(request.form["id_offerta"])])
                mcpDB.commit()
                mcpDB.close()
                flash("File eliminato con successo!", "success")
            except:
                flash("Errore nel caricamento del file!", "warning")
        elif request.form["id_form"] == "modifica_offerta":
            try:
                anno, mese, giorno = request.form["data"].split('-')
                tmp = datetime.date(int(anno), int(mese), int(giorno))
                dati_validi = True
            except:
                dati_validi = False
                flash("Dati errati: Manca la data di creazione offerta!", "warning")
            tmp_date = datetime.datetime.strptime(request.form["data"], "%Y-%m-%d").date()
            if tmp_date > datetime.datetime.now().date():
                flash("Ehi stai viaggiando nel tempo?", "warning")
                dati_validi = False
            tmp_id_cliente = request.form["cliente"]
            if request.form["nome"] == "":
                dati_validi = False
                flash("Dati errati: Non puoi creare un'offerta senza ID!", "warning")
            if request.form["n_ordine"] == "":
                n_ordine = ""
            else:
                n_ordine = request.form["n_ordine"]
            tmp_dati = {
                "form": request.form["id_form"],
                "id_offerta": request.form["id_offerta"],
                "nome": request.form["nome"],
                "cliente": tmp_id_cliente,
                "n_ordine": n_ordine,
                "data": request.form["data"],
                "note": request.form["note"]
                }
            if dati_validi:
                insert_edit_offerta(tmp_dati)
                flash("Inserito con successo!", "success")
    mcpDB = sqlite3.connect(mcpDB_path)
    cur = mcpDB.cursor()
    cur.execute("select * from offerte where id = ?", [id_offerta])
    offerte = cur.fetchall()
    for offerta in offerte:
        tmp_versioni = []
        for i in json.loads(offerta[4])["versioni"]:
            tmp_versioni.append({"rev": i["rev"], "file": {"id": i["file"],"nome": get_nome("file", i["file"])}})
        tmp_offerta = {
            "id": offerta[0],
            "id_numero": offerta[1],
            "numero": get_numero_offerta(offerta[1], json.loads(offerta[4])),
            "ordine_rif": offerta[2],
            "cliente": {"id": offerta[3],"nome": get_nome("cliente", offerta[3])},
            "versioni": {"data": json.loads(offerta[4])["data"], "chiusa": json.loads(offerta[4])["chiusa"], "versioni": tmp_versioni},
            "note": offerta[5]
            }
    mcpDB.commit()
    mcpDB.close()
    return render_template("dettaglio_offerta.html", offerta=tmp_offerta, gestione=gestisce(), clienti=get_clienti(), utenti=User.query.all(), pre_suff=ct["offerte"], menu_page="offerte")

@app.route("/dettaglio_commessa_interna/<string:id_commessa>")
@login_required
def dettaglio_storico_interna(id_commessa):
    if not gestisce():
        return redirect(url_for("dashboard"))
    mcpDB = sqlite3.connect(mcpDB_path)
    cur = mcpDB.cursor()
    cur.execute("select * from interne where id = ?", [int(id_commessa)])
    commesse = cur.fetchall()
    cur.execute("select * from ore_interne where id_interna = ?", [int(id_commessa)])
    lavorazioni = cur.fetchall()
    mcpDB.commit()
    mcpDB.close()
    tmp_lavorazioni = []
    ore_lavorate = 0.0
    euro_spesi = 0.0
    for lavorazione in lavorazioni:
        tmp_lavorazione = {
            "id": lavorazione[0],
            "collaboratore": {"id": lavorazione[1],"nome": get_nome("collaboratore", lavorazione[1])},
            "durata": json.loads(lavorazione[3]),
            "costo": get_costo(lavorazione[1], json.loads(lavorazione[3]))
            }
        ore_lavorate += float(tmp_lavorazione["durata"]["ore"])
        euro_spesi += tmp_lavorazione["costo"]
        tmp_lavorazioni.append(tmp_lavorazione)
    tmp_lavorazioni.sort(key = lambda x: datetime.datetime.strptime(x["durata"]["data"], '%Y-%m-%d'))
    tmp_specifiche = {
            "ore_lavorate": ore_lavorate,
            "euro_spesi": euro_spesi
            }
    tmp_commessa = {
        "id": commesse[0][0],
        "numero": get_numero_commessa(commesse[0][1]),
        "nome": commesse[0][2],
        "specifiche": tmp_specifiche,
        "durata": json.loads(commesse[0][3]),
        "note": commesse[0][4],
        "lavorazioni": tmp_lavorazioni
        }
    return render_template("dettaglio_storico_interna.html", commessa=tmp_commessa, gestione=gestisce(), clienti=get_clienti(), utenti=User.query.all(), tipi_lavorazione=get_tipo_lavorazione(), menu_page="storico_commesse_interne")

@app.route("/storico", defaults={ "anno": False })
@app.route("/storico/<string:anno>")
@login_required
def storico(anno):
    if not gestisce():
        return redirect(url_for("dashboard"))
    return render_template("storico.html", anno=anno, anni=anni_commesse(), gestione=gestisce(), dati=dati_mcp("commesse_complete"), clienti=get_clienti(), utenti=User.query.all(), tipi_lavorazione=get_tipo_lavorazione(), menu_page="storico_commesse")

@app.route("/storico_interne", defaults={ "anno": False })
@app.route("/storico_interne/<string:anno>")
@login_required
def storico_interne(anno):
    if not gestisce():
        return redirect(url_for("dashboard"))
    return render_template("storico_interne.html", anno=anno, anni=anni_commesse(), gestione=gestisce(), dati=dati_mcp("interne"), clienti=get_clienti(), utenti=User.query.all(), tipi_lavorazione=get_tipo_lavorazione(), menu_page="storico_commesse_interne")

@app.route("/clienti", methods=("GET", "POST"))
@login_required
def clienti():
    if not gestisce():
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        if request.form["id_form"] == "nuovo_cliente":
            mcpDB = sqlite3.connect(mcpDB_path)
            cur = mcpDB.cursor()
            cur.execute("INSERT INTO clienti (nome) VALUES (?)", [request.form["nome"]])
            mcpDB.commit()
            mcpDB.close()
        elif request.form["id_form"] == "modifica_cliente":
            mcpDB = sqlite3.connect(mcpDB_path)
            cur = mcpDB.cursor()
            cur.execute("UPDATE clienti SET nome = ? WHERE id = ?", [request.form["nome"], int(request.form["id_cliente"])])
            mcpDB.commit()
            mcpDB.close()
        elif request.form["id_form"] == "elimina_cliente":
            commesse = dati_mcp("commesse")
            ha_commesse = False
            for i in commesse:
                if int(i["cliente"]["id"]) == int(request.form["id_cliente"]):
                    ha_commesse = True
            if ha_commesse:
                flash("Cliente NON eliminato, esistono commesse a suo nome!", "warning")
            else:
                mcpDB = sqlite3.connect(mcpDB_path)
                cur = mcpDB.cursor()
                cur.execute("DELETE from clienti where id = ?", [int(request.form["id_cliente"])])
                mcpDB.commit()
                mcpDB.close()
                flash("Eliminato con successo!", "success")
    return render_template("clienti.html", gestione=gestisce(), clienti=get_clienti(), utenti=User.query.all(), tipi_lavorazione=get_tipo_lavorazione(), menu_page="database")

@app.route("/tipologie", methods=("GET", "POST"))
@login_required
def tipologie():
    if not gestisce():
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        if request.form["id_form"] == "inserisci_tipologia":
            tmp_tag = request.form["nome"].lower().replace(" ", "_")
            mcpDB = sqlite3.connect(mcpDB_path)
            cur = mcpDB.cursor()
            cur.execute("INSERT INTO tipo_lavorazioni (tag, nome) VALUES (?,?)", [tmp_tag, request.form["nome"]])
            mcpDB.commit()
            mcpDB.close()
    return render_template("tipologie.html", gestione=gestisce(), clienti=get_clienti(), utenti=User.query.all(), tipi_lavorazione=get_tipo_lavorazione(), menu_page="database")

@app.route("/report", methods=("GET", "POST"))
@login_required
def report():
    if not gestisce():
        return redirect(url_for("dashboard"))
    base_anno = 2020
    oggi = [int(str(datetime.datetime.now())[0:4]), int(str(datetime.datetime.now())[5:7])]
    mesi = []
    anni = []
    for i in range(oggi[0] - base_anno+1):
        anni.append(i+base_anno)
        for j in range(12):
            mesi.append([j+1, i+base_anno])
    for i in range(oggi[1]-1):
        mesi.append([i+1, 2023])
    anni.reverse()
    if request.method == "POST":
        if request.form["id_form"] == "report_settimanale":
            return redirect(url_for("report_periodo", tipo=request.form["tipo"], anno=datetime.datetime.strptime(request.form["settimana"] + '-1', "%Y-W%W-%w").year, mese=datetime.datetime.strptime(request.form["settimana"] + '-1', "%Y-W%W-%w").month, settimana=request.form["settimana"]))
        elif request.form["id_form"] == "report_mensile":
            return redirect(url_for("report_periodo", tipo=request.form["tipo"], anno=request.form["mese"][0:4], mese=request.form["mese"][5:7], settimana="0"))
        elif request.form["id_form"] == "report_annuale" and int(request.form["anno"]) in anni:
            return redirect(url_for("report_periodo", tipo=request.form["tipo"], anno=request.form["anno"], mese="0", settimana="0"))
    return render_template("report.html", anni=anni, gestione=gestisce(), utenti=User.query.all(), menu_page="report")

@app.route("/report/<string:tipo>/<string:anno>/<string:mese>/<string:settimana>")
@login_required
def report_periodo(tipo, anno, mese, settimana):
    if not gestisce():
        return redirect(url_for("dashboard"))
    if tipo == "commesse":
        dati = dati_mcp("commesse_complete")
        commesse = []
        for i in dati:
            tmp_bool = False
            for j in i["lavorazioni"]:
                if int(j["durata"]["data"][0:4]) == int(anno):
                    if int(mese) == 0:
                        tmp_bool = True
                    elif int(j["durata"]["data"][5:7]) == int(mese):
                        if settimana == "0":
                            tmp_bool = True
                if settimana != "0":
                    for y in range(7):
                        giorno = datetime.datetime.strptime(settimana + '-1', "%Y-W%W-%w")+datetime.timedelta(days=y)
                        if int(j["durata"]["data"][8:10]) == int(giorno.day) and int(j["durata"]["data"][5:7]) == int(giorno.month) and int(j["durata"]["data"][0:4]) == int(giorno.year):
                            tmp_bool = True
            if tmp_bool:
                commesse.append(i)
        workbook = load_workbook(filename = static_path+'report_commesse.xlsx')
        for i in commesse:
            workbook.copy_worksheet(workbook['base']).title = i["numero"]
            foglio = workbook[i["numero"]]
            foglio["B1"] = i["numero"]
            foglio["B2"] = i["nome"]
            foglio["B3"] = i["cliente"]["nome"]
            foglio["B4"] = i["durata"]["inizio"][8:10]+"/"+i["durata"]["inizio"][5:7]+"/"+i["durata"]["inizio"][0:4]
            if i["durata"]["fine"]:
                foglio["B5"] = i["durata"]["fine"][8:10]+"/"+i["durata"]["fine"][5:7]+"/"+i["durata"]["fine"][0:4]
            foglio["B6"] = i["specifiche"]["project_manager"]["nome"]
            foglio["B7"] = i["specifiche"]["budget_ore"]
            foglio["B8"] = i["specifiche"]["budget_euro"]

            tmp_riga = 12
            for j in i["lavorazioni"]:
                if int(j["durata"]["data"][0:4]) == int(anno):
                    if int(mese) == 0 or (int(j["durata"]["data"][5:7]) == int(mese)):
                        foglio.cell(row=tmp_riga, column=1).value = j["durata"]["data"][8:10]+"/"+j["durata"]["data"][5:7]+"/"+j["durata"]["data"][0:4]
                        foglio.cell(row=tmp_riga, column=2).value = j["collaboratore"]["nome"]
                        foglio.cell(row=tmp_riga, column=3).value = float(j["durata"]["ore"])
                        if j["durata"]["trasferta"]:
                            foglio.cell(row=tmp_riga, column=4).value = "Si"
                        else:
                            foglio.cell(row=tmp_riga, column=4).value = "No"
                        foglio.cell(row=tmp_riga, column=5).value = j["tipo"]["nome"]
                        foglio.cell(row=tmp_riga, column=6).value = j["costo"]
                        tmp_riga += 1
        try:
            workbook.remove(workbook['base'])
            workbook.save(filename=static_path+"report_commesse_compilato.xlsx")
        except:
            flash("Non ci sono commesse nel periodo selezionato!", "warning")
            return redirect(url_for("report"))
        if int(mese) == 0:
            nome_file = "report_commesse_"+anno+".xlsx"
        elif settimana == "0":
            nome_file = "report_commesse_"+mese+"_"+anno+".xlsx"
        else:
            nome_file = "report_commesse_"+settimana+".xlsx"

        return send_file(static_path+"report_commesse_compilato.xlsx", as_attachment=True, download_name=nome_file)
    elif tipo == "collaboratori":
        mcpDB = sqlite3.connect(mcpDB_path)
        cur = mcpDB.cursor()
        cur.execute("select * from lavorazioni")
        lavorazioni = cur.fetchall()
        tmp_lavorazioni = []
        for lavorazione in lavorazioni:
            tmp_lavorazione = {
                "id": lavorazione[0],
                "collaboratore": {"id": lavorazione[1],"nome": get_nome("collaboratore", lavorazione[1])},
                "commessa" : lavorazione[2],
                "tipo": {"id": lavorazione[3],"nome": get_nome("tipo_lavorazione", lavorazione[3])},
                "durata": json.loads(lavorazione[4]),
                "tipologia": "lavorazione",
                }
            if int(mese) == 0 and int(tmp_lavorazione["durata"]["data"][0:4]) == int(anno):
                tmp_lavorazioni.append(tmp_lavorazione)
            elif int(tmp_lavorazione["durata"]["data"][5:7]) == int(mese) and int(tmp_lavorazione["durata"]["data"][0:4]) == int(anno) and settimana == "0":
                tmp_lavorazioni.append(tmp_lavorazione)
            elif settimana != "0":
                for y in range(7):
                    giorno = datetime.datetime.strptime(settimana + '-1', "%Y-W%W-%w")+datetime.timedelta(days=y)
                    if int(tmp_lavorazione["durata"]["data"][8:10]) == int(giorno.day) and int(tmp_lavorazione["durata"]["data"][5:7]) == int(giorno.month) and int(tmp_lavorazione["durata"]["data"][0:4]) == int(giorno.year):
                        tmp_lavorazioni.append(tmp_lavorazione)

        cur.execute("select * from ore_interne")
        lavorazioni = cur.fetchall()
        for lavorazione in lavorazioni:
            tmp_lavorazione = {
                "id": lavorazione[0],
                "collaboratore": {"id": lavorazione[1],"nome": get_nome("collaboratore", lavorazione[1])},
                "commessa" : lavorazione[2],
                "durata": json.loads(lavorazione[3]),
                "tipologia": "interne",
                }
            if int(mese) == 0 and int(tmp_lavorazione["durata"]["data"][0:4]) == int(anno):
                tmp_lavorazioni.append(tmp_lavorazione)
            elif int(tmp_lavorazione["durata"]["data"][5:7]) == int(mese) and int(tmp_lavorazione["durata"]["data"][0:4]) == int(anno) and settimana == "0":
                tmp_lavorazioni.append(tmp_lavorazione)
            elif settimana != "0":
                for y in range(7):
                    giorno = datetime.datetime.strptime(settimana + '-1', "%Y-W%W-%w")+datetime.timedelta(days=y)
                    if int(tmp_lavorazione["durata"]["data"][8:10]) == int(giorno.day) and int(tmp_lavorazione["durata"]["data"][5:7]) == int(giorno.month) and int(tmp_lavorazione["durata"]["data"][0:4]) == int(giorno.year):
                        tmp_lavorazioni.append(tmp_lavorazione)
        tmp_lavorazioni.sort(key = lambda x: datetime.datetime.strptime(x["durata"]["data"], '%Y-%m-%d'))
        utenti = User.query.all()
        workbook = load_workbook(filename = static_path+'report_collaboratori.xlsx')
        for i in utenti:
            workbook.copy_worksheet(workbook['base']).title = i.anagrafica["nome"] + " " + i.anagrafica["cognome"]
            foglio = workbook[i.anagrafica["nome"] + " " + i.anagrafica["cognome"]]

            tmp_riga = 2
            for j in tmp_lavorazioni:
                if int(j["collaboratore"]["id"]) == int(i.id):
                    if j["tipologia"] == "lavorazione":
                        cur.execute("select * from commesse where id = ?", [j["commessa"]])
                        commesse = cur.fetchall()
                        commessa = commesse[0]
                        tmp_specifiche = {
                                "budget_ore": json.loads(commessa[4])["budget_ore"],
                                "budget_euro": json.loads(commessa[4])["budget_euro"],
                                "project_manager": {"id": json.loads(commessa[4])["project_manager"],"nome": get_nome("collaboratore", json.loads(commessa[4])["project_manager"])},
                                "tipologia": get_tipo_commessa(json.loads(commessa[4])["tipologia"])
                                }
                        tmp_commessa = {
                            "id": commessa[0],
                            "numero": get_numero_commessa(commessa[1]),
                            "nome": commessa[2],
                            "cliente": {"id": commessa[3],"nome": get_nome("cliente", commessa[3])},
                            "specifiche": tmp_specifiche,
                            "durata": json.loads(commessa[5]),
                            "note": commessa[6]
                            }
                        foglio.cell(row=tmp_riga, column=1).value = j["durata"]["data"][8:10]+"/"+j["durata"]["data"][5:7]+"/"+j["durata"]["data"][0:4]
                        foglio.cell(row=tmp_riga, column=2).value = tmp_commessa["numero"]
                        foglio.cell(row=tmp_riga, column=3).value = tmp_commessa["nome"]
                        foglio.cell(row=tmp_riga, column=4).value = j["tipo"]["nome"]
                        foglio.cell(row=tmp_riga, column=5).value = float(j["durata"]["ore"])
                        if j["durata"]["trasferta"]:
                            foglio.cell(row=tmp_riga, column=6).value = "Si"
                        else:
                            foglio.cell(row=tmp_riga, column=6).value = "No"
                        tmp_riga += 1
                    elif j["tipologia"] == "interne":
                        cur.execute("select * from interne where id = ?", [j["commessa"]])
                        commesse = cur.fetchall()
                        commessa = commesse[0]
                        tmp_commessa = {
                            "id": commessa[0],
                            "numero": get_numero_commessa(commessa[1]),
                            "nome": commessa[2],
                            "durata": json.loads(commessa[3]),
                            "note": commessa[4]
                            }
                        foglio.cell(row=tmp_riga, column=1).value = j["durata"]["data"][8:10]+"/"+j["durata"]["data"][5:7]+"/"+j["durata"]["data"][0:4]
                        foglio.cell(row=tmp_riga, column=2).value = tmp_commessa["numero"]
                        foglio.cell(row=tmp_riga, column=3).value = tmp_commessa["nome"]
                        foglio.cell(row=tmp_riga, column=5).value = float(j["durata"]["ore"])
                        if j["durata"]["trasferta"]:
                            foglio.cell(row=tmp_riga, column=6).value = "Si"
                        else:
                            foglio.cell(row=tmp_riga, column=6).value = "No"
                        tmp_riga += 1
        mcpDB.commit()
        mcpDB.close()
        try:
            workbook.remove(workbook['base'])
            workbook.save(filename=static_path+"report_collaboratori_compilato.xlsx")
        except:
            flash("Qualcosa non ha funzionato!", "warning")
            return redirect(url_for("report"))
        if int(mese) == 0:
            nome_file = "report_collaboratori_"+anno+".xlsx"
        elif settimana == "0":
            nome_file = "report_collaboratori_"+mese+"_"+anno+".xlsx"
        else:
            nome_file = "report_collaboratori_"+settimana+".xlsx"

        return send_file(static_path+"report_collaboratori_compilato.xlsx", as_attachment=True, download_name=nome_file)
    return redirect(url_for("report"))

@app.route("/genera_report/<string:id_commessa>")
@login_required
def crea_report(id_commessa):
    if not gestisce():
        return redirect(url_for("dashboard"))
    mcpDB = sqlite3.connect(mcpDB_path)
    cur = mcpDB.cursor()
    cur.execute("select * from commesse where id = ?", [int(id_commessa)])
    commesse = cur.fetchall()
    cur.execute("select * from lavorazioni where commessa = ?", [int(id_commessa)])
    lavorazioni = cur.fetchall()
    mcpDB.commit()
    mcpDB.close()
    tmp_lavorazioni = []
    ore_lavorate = 0.0
    euro_spesi = 0.0
    for lavorazione in lavorazioni:
        tmp_lavorazione = {
            "id": lavorazione[0],
            "collaboratore": {"id": lavorazione[1],"nome": get_nome("collaboratore", lavorazione[1])},
            "tipo": {"id": lavorazione[3],"nome": get_nome("tipo_lavorazione", lavorazione[3])},
            "durata": json.loads(lavorazione[4]),
            "costo": get_costo(lavorazione[1], json.loads(lavorazione[4]))
            }
        ore_lavorate += float(tmp_lavorazione["durata"]["ore"])
        euro_spesi += tmp_lavorazione["costo"]
        tmp_lavorazioni.append(tmp_lavorazione)
    tmp_lavorazioni.sort(key = lambda x: datetime.datetime.strptime(x["durata"]["data"], '%Y-%m-%d'))
    tmp_specifiche = {
            "budget_ore": float(json.loads(commesse[0][4])["budget_ore"]),
            "budget_euro": float(json.loads(commesse[0][4])["budget_euro"]),
            "ore_lavorate": ore_lavorate,
            "euro_spesi": euro_spesi,
            "project_manager": {"id": json.loads(commesse[0][4])["project_manager"],"nome": get_nome("collaboratore", json.loads(commesse[0][4])["project_manager"])},
            "tipologia": get_tipo_commessa(json.loads(commesse[0][4])["tipologia"])
            }
    tmp_commessa = {
        "id": commesse[0][0],
        "numero": get_numero_commessa(commesse[0][1]),
        "nome": commesse[0][2],
        "cliente": {"id": commesse[0][3],"nome": get_nome("cliente", commesse[0][3])},
        "specifiche": tmp_specifiche,
        "durata": json.loads(commesse[0][5]),
        "note": commesse[0][6]
        }
    workbook = load_workbook(filename = static_path+'report_commessa.xlsx')
    foglio = workbook["Riepilogo"]

    foglio["B1"] = tmp_commessa["numero"]
    foglio["B2"] = tmp_commessa["nome"]
    foglio["B3"] = tmp_commessa["cliente"]["nome"]
    foglio["B4"] = tmp_commessa["durata"]["inizio"][8:10]+"/"+tmp_commessa["durata"]["inizio"][5:7]+"/"+tmp_commessa["durata"]["inizio"][0:4]
    foglio["B5"] = tmp_commessa["durata"]["fine"][8:10]+"/"+tmp_commessa["durata"]["fine"][5:7]+"/"+tmp_commessa["durata"]["fine"][0:4]
    foglio["B6"] = tmp_commessa["specifiche"]["project_manager"]["nome"]
    foglio["B7"] = tmp_commessa["specifiche"]["budget_ore"]
    foglio["B8"] = tmp_commessa["specifiche"]["budget_euro"]

    tmp_riga = 12
    for i in tmp_lavorazioni:
        foglio.cell(row=tmp_riga, column=1).value = i["durata"]["data"][8:10]+"/"+i["durata"]["data"][5:7]+"/"+i["durata"]["data"][0:4]
        foglio.cell(row=tmp_riga, column=2).value = i["collaboratore"]["nome"]
        foglio.cell(row=tmp_riga, column=3).value = float(i["durata"]["ore"])
        if i["durata"]["trasferta"]:
            foglio.cell(row=tmp_riga, column=4).value = "Si"
        else:
            foglio.cell(row=tmp_riga, column=4).value = "No"
        foglio.cell(row=tmp_riga, column=5).value = i["tipo"]["nome"]
        foglio.cell(row=tmp_riga, column=6).value = i["costo"]
        tmp_riga += 1

    workbook.save(filename=static_path+"report_commessa_compilato.xlsx")

    nome_file = "report_commessa_"+tmp_commessa["durata"]["inizio"][0:4]+"_"+tmp_commessa["numero"]+".xlsx"

    return send_file(static_path+"report_commessa_compilato.xlsx", as_attachment=True, download_name=nome_file)

#GESTIONE ACCOUNT - GESTIONE LOGIN
@app.route("/account", methods=("GET", "POST"))
@login_required
def account():
    if request.method == "POST":
        if request.form["id_form"] == "aggiorna_passwd":
            if check_password_hash(current_user.password, request.form["old_passwd"]) and request.form["passwd"] == request.form["new_passwd"] and request.form["passwd"] != "":
                tmp_password = generate_password_hash(request.form["passwd"])
                utente = User.query.filter_by(username=current_user.username).first()
                utente.password = tmp_password
                db.session.commit()
                flash("Password modificata con successo!", "success")
            else:
                flash("Sicuro di aver scritto correttamente le password?", "warning")
    return render_template("account.html", gestione=gestisce(), menu_page="account")

@app.route("/login", methods=("GET", "POST"))
def login():
    if request.method == "POST":
        utente = User.query.filter_by(username=request.form["username"]).first()
        if utente:
            if check_password_hash(utente.password, request.form["passwd"]):
                login_user(utente)
                return redirect(url_for("homepage"))
    return render_template("login.html")

@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("homepage"))

@app.route("/collaboratori", methods=("GET", "POST"))
@login_required
def collaboratori():
    if not gestisce():
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        if request.form["id_form"] == "nuovo_utente":
            utente = User.query.filter_by(username=request.form["username"]).first()
            if utente is None:
                password = generate_password_hash(request.form["passwd"])
                tmp_cod = random.randint(111111,999999)
                anagrafica = {
                    "nome": request.form["nome"],
                    "cognome": request.form["cognome"],
                    "mail": request.form["mail"]
                    }
                verifica = {
                    "mail": tmp_cod,
                    "recupero": ""
                    }
                collaboratore = {
                    "ruolo": request.form["ruolo"],
                    "costo": request.form["costo"] if request.form["costo"] != "" else "0"
                    }

                utente = User(username=request.form["username"], password=password, anagrafica=anagrafica, collaboratore=collaboratore, verifica=verifica)
                db.session.add(utente)
                db.session.commit()
                flash("Utente inserito con successo!", "success")
            else:
                flash("Esiste già l'utente "+request.form["username!"], "warning")
        elif request.form["id_form"] == "modifica_utente":
            utente = User.query.filter_by(id=int(request.form["id_collaboratore"])).first()
            anagrafica = {
                "nome": request.form["nome"],
                "cognome": request.form["cognome"],
                "mail": request.form["mail"]
                }
            collaboratore = {
                "ruolo": request.form["ruolo"],
                "costo": request.form["costo"] if request.form["costo"] != "" else "0"
                }
            utente.username = request.form["username"]
            utente.anagrafica = anagrafica
            utente.collaboratore = collaboratore
            db.session.commit()
            flash("Utente modificato con successo!", "success")
        elif request.form["id_form"] == "modifica_passwd":
            utente = User.query.filter_by(username=request.form["username"]).first()
            tmp_password = generate_password_hash(request.form["passwd"])
            utente.password = tmp_password
            db.session.commit()
    utenti = []
    for i in User.query.all():
        utente = {
            "id": i.id,
            "username": i.username,
            "anagrafica": i.anagrafica,
            "collaboratore": i.collaboratore
            }
        utenti.append(utente)
    return render_template("collaboratori.html", gestione=gestisce(), utenti=utenti, menu_page="database")

@app.route("/download/<string:id_file>")
@login_required
def download(id_file):
    mcpDB = sqlite3.connect(mcpDB_path)
    cur = mcpDB.cursor()
    cur.execute("select * from file where id = ?", [id_file])
    file = cur.fetchall()
    mcpDB.commit()
    mcpDB.close()
    return send_file(BytesIO(file[0][2]), download_name=file[0][1], as_attachment=True)

@app.errorhandler(404)
def page_not_found(e):
    return render_template("404.html"), 404

@app.errorhandler(405)
def internal_error(e):
    return render_template("errore_generico.html"), 405

@app.errorhandler(500)
def internal_error(e):
    return render_template("errore_generico.html"), 500
